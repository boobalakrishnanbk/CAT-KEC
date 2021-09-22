from django.db.models.expressions import Exists
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from cat_app.models import Mark

import openpyxl
from openpyxl import Workbook

# Create your views here.
@login_required(login_url='/login/')
def staff(request):
    error = False
    criteria_error = False
    subjects_hide = True
    subjects = []
    students_details = []
    if request.method == 'POST':
        result = Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).values('roll_number').distinct()
        students_roll_number = result
        students_details = {}
        subjects = Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).values('subject_name').distinct()
    
        criteria = {}
        count = 0
        for i in subjects:
            criteria[i["subject_name"]] = count
            count += 1
        
        if request.POST['file_type'] == "mark" and request.POST:
            mark_file = request.FILES['mark_file']
            if Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).count()==0:
                error = importMark(mark_file, request)
            else:
                Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat']).delete()
                error = importMark(mark_file, request)
        if request.POST['file_type'] == "marks":
            for student in students_roll_number:
                result = Mark.objects.filter(semester=request.POST['semester'],cat=request.POST['cat'],roll_number = student['roll_number']).values().distinct()

                detail = []
                detail.append(["name",result[0]['name']])
                count = 0
                t = {}
                for i in criteria:
                    if str(i) not in t.keys():
                        t[i] = "-"
                for mark in result:
                    t[mark['subject_name']] =mark['mark']
                detail.append(t)
                students_details[student['roll_number']] = detail        
            if students_roll_number.count() > 0:
                criteria_error = False
                subjects_hide = False
            else:
                subjects_hide = True
                criteria_error = "Sorry! The selected criteria result was not yet uploaded. Please Choose other option"
    
    data = {'semester':[],'cat':[]}
    mark =  Mark.objects.all()
    data['semester'] = mark.values('semester').distinct().order_by('semester')
    data['cat'] = mark.values('cat').distinct().order_by('cat')
    return render(
        request, 
        'staff_page.html', 
        {
            "error":error,
            "c_error":criteria_error,
            "subjects":subjects,
            "marks":students_details,
            "data": data,
            "display":subjects_hide
            
        }
    )


# import to database - Mark
def importMark(mark_file, request):
    try:
        workbook = openpyxl.load_workbook(mark_file, read_only=True)
        first_sheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(first_sheet)
        marks = []
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            roll_no = row[0].value
            name = row[1].value
            phone = row[2].value
            try:
                phone = phone.replace(" ","")
            except:
                pass
            for i in range(4, worksheet.max_column,2):
                mark = Mark()
                mark.roll_number = roll_no
                mark.name = name
                mark.phone = phone
                mark.subject_name = row[i].value
                mark.mark = row[i+1].value
                mark.semester = request.POST['semester']
                mark.cat = request.POST['cat']
                marks.append(mark)
        
        if Mark.objects.bulk_create(marks):
            cat = request.POST['cat']=="End Semester Exam" and "End Semester Examination" or ("CAT - "+request.POST['cat'])
            return "Successfully uploaded the marks for "  + " semsester " + request.POST['semester']+ " " + cat
        else:    
            return "Invalid Excel Format"
    except:    
        return "Invalid Excel Format"
    
# homepage
def home(request):
    if request.user.is_authenticated:
        return redirect('/staff/')
    return render(request, 'home.html', {})



def studentLogin(request):
    data = {'semester':[],'cat':[]}
    mark =  Mark.objects.filter(roll_number = request.POST['roll_number'].upper())
    data['semester'] = mark.values('semester').distinct().order_by('semester')
    data['cat'] = mark.values('cat').distinct().order_by('cat')

    if request.POST['phone'] and request.POST['roll_number']:
        marks = Mark.objects.filter(roll_number=request.POST['roll_number'].upper().strip(),phone=request.POST['phone'].strip()).count()
        if marks>1:
            request.session['roll_number'] = request.POST['roll_number']
            request.session['phone'] = request.POST['phone']
            return render(request, 'show.html', {"show":"d-none",'data':data})
    return render(request, 'home.html', {'error':True,'data':data})

def fetch_marks(request):
    att = ""
    if request.POST['semester'] and request.POST['cat']:
        marks = Mark.objects.filter(roll_number=request.session['roll_number'].upper(),phone=request.session['phone'],semester=request.POST['semester'],cat=request.POST['cat'])
        if marks.count()>0:
            gpa= 0
            cgpa = 0
            subs_marks = {}
            try:
                int(marks.values("cat").distinct()[0]['cat'])
                for i in marks:
                    if not i.subject_name == "ATT":
                        if not i.mark == None:
                            remark = []
                            remark.append(int(i.mark))
                            if int(i.mark)>=50:remark.append("Pass")
                            else: remark.append("Fail")
                            subs_marks[i.subject_name] = remark
                    else:
                            att = i.mark
            except ValueError:
                cgpa = 0
                gpa= 0
                for i in marks:
                    if i.subject_name == "Gpa":
                        if not i.mark == None:
                            gpa = i.mark
                            
                    elif i.subject_name == "Cgpa":
                        if not i.mark == None:
                            cgpa = i.mark
                            
                    else:
                        if not i.mark == None:
                            remark = []
                            remark.append(i.mark)
                            passed = ['S','A','B','C','D','E','O','A+','A','B','B+']
                            if i.mark in passed:
                                remark.append("Pass")
                            else:
                                remark.append("Fail")
                            subs_marks[i.subject_name] = remark
            exam = marks.values("cat").distinct()[0]
            
            data = {'semester':[],'cat':[]}
            mark =  Mark.objects.filter(roll_number=request.session['roll_number'].upper())
            data['semester'] = mark.values('semester').distinct().order_by('semester')
            data['cat'] = mark.values('cat').distinct().order_by('cat')

            return render(request, 'show.html', {
                "roll_number":request.session['roll_number'],
                "show":"",
                "name":marks.values("name").distinct()[0],
                "semester":marks.values("semester").distinct()[0],
                "exam": exam['cat']=="End Semester Exam" and "End Semester Examination" or ("CAT - "+str(marks.values("cat").distinct()[0]['cat'])),
                "sub_marks":subs_marks,
                "attendance":att,
                "data":data,
                "gpa":gpa,
                "cgpa":cgpa,
                "error":False
            })

        else:
            data = {'semester':[],'cat':[]}
            mark =  Mark.objects.filter(roll_number=request.session['roll_number'].upper())
            data['semester'] = mark.values('semester').distinct().order_by('semester')
            data['cat'] = mark.values('cat').distinct().order_by('cat')
            return render(request, 'show.html', {
                "roll_number":request.session['roll_number'],
                "show":"d-none",
                'data':data,
                'error':"Sorry! The selected result was not yet uploaded. Please Choose other option",
            })

    

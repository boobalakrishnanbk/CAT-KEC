
import openpyxl
from openpyxl import Workbook
import os

# import formattinf sheet
student_detail_file = "/var/www/html/project/result-20210916T132001Z-001/result/mark(2020).xlsx"
workbook = openpyxl.load_workbook(student_detail_file)
first_sheet = workbook.get_sheet_names()[0]
old_file = workbook.get_sheet_by_name(first_sheet)

#create neww sheet format
wb = openpyxl.Workbook()
sheet = wb.active

# create headings for row
cell = sheet.cell(row = 1, column = 1)
cell.value = "REG_NO"
cell = sheet.cell(row = 1, column = 2)
cell.value = "Name"
cell = sheet.cell(row = 1, column = 3)
cell.value = "Phone"
cell = sheet.cell(row = 1, column = 4)
cell.value = "Exam"

temp_count = 1
for i in range(5,35,2):
    cell = sheet.cell(row = 1, column = i)
    cell.value = "Subject - " + str(temp_count)
    
    cell = sheet.cell(row = 1, column = i+1)
    cell.value = "Mark - " + str(temp_count)
    temp_count += 1
    
# creating old data to new formated data
temp_row_count = 2
temp_column_count = 0

for row in old_file.iter_rows(min_row=2, max_row=old_file.max_row):
    # name
    cell = sheet.cell(row = temp_row_count, column = 1)
    cell.value = row[1].value
    # roll number
    cell = sheet.cell(row = temp_row_count, column = 2)
    cell.value = row[0].value
    # exam
    cell = sheet.cell(row = temp_row_count, column = 4)
    cell.value = "End Semester"

    temp_column_count = 5
    count = 2
    sam = 0
    for col in range(2,old_file.max_column):
        if not row[col].value == "-":
            # subject - count
            cell = sheet.cell(row = temp_row_count, column = temp_column_count)
            title = (old_file[1])
            cell.value = title[col].value
            temp_column_count += 1
            # mark - count
            cell = sheet.cell(row = temp_row_count, column = temp_column_count)
            cell.value = row[col].value
            temp_column_count += 1
    temp_row_count += 1

# final creating excel sheet
wb.save("/var/www/html/project/result-20210916T132001Z-001/result/mark(2020)_re_result.xlsx")
